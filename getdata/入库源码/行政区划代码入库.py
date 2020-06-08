#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time    : 2018/6/26 14:40
# @Author  : legend

from openpyxl import load_workbook
import pymssql
# wb = load_workbook("C:\\Users\\legend\\Desktop\\工作簿1.xlsx")
# # sheet = wb.get_sheet_by_name("Sheet1")网上找的资源，get_sheet_by_name函数被弃用了。直接wb[sheetname]
# sheet = wb["Sheet1"]
# print(sheet["A4"].value)
class MSSQL_Interface :
    # 数据库信息
    def __init__(self,host,user,pwd,dbname,table_name):
        self.host = host
        self.user = user
        self.pwd = pwd
        self.dbname = dbname
        self.connect_db()
        self.table_name = table_name

    # 链接数据库
    def connect_db(self):
        try:
            self.conn = pymssql.connect(user=self.user,password=self.pwd,host=self.host,database=self.dbname,charset="utf8")
            self.cur = self.conn.cursor()
        except:
            print("ERROR:fail to connect mssql")
    def __del__(self):
        self.cur.close()
        self.conn.close()
    def read_xlsx(self,xlsx_path):
        wb = load_workbook(xlsx_path)
        # sheet = wb.get_sheet_by_name("Sheet1")网上找的资源，get_sheet_by_name函数被弃用了。直接wb[sheetname]
        sheet = wb["Sheet1"]
        date_list = []
        for i in range(3, sheet.max_row + 1):
            temp_list = []
            code = sheet["A" + str(i)].value
            area_name = sheet["B" + str(i)].value.strip()
            temp_list.append(code)
            temp_list.append(area_name)
            for i in range(0, 9):
                temp_list.append('')
            temp_list.append("0")
            date_list.append(tuple(temp_list))
        return date_list
    def import_list(self, data_list, table_name):  # 输入已经整理好的list格式
        q0 = "use " + self.dbname + ';'
        q1 = "insert into " + '[' +table_name+ ']' + " values("
        temp_q = ""
        for i in range(0, data_list[0].__len__()):
            temp_q = temp_q + "%s"
            if i < data_list[0].__len__() - 1:
                temp_q = temp_q + ','
        query = q0 + q1 + temp_q + ")"
        try:
            self.cur.executemany(query,data_list)
            self.conn.commit()
        except Exception as e:
            print("执行sql: % s时出错： % s" % (query, e))

    def import_file(self,xlsx_path):
        date = self.read_xlsx(xlsx_path)
        self.import_list(date,self.table_name)
if __name__ == '__main__':
    host = "localhost"
    user = "sa"
    pwd = "123456"
    db = "address_model"
    table_name = "行政区划"
    item = MSSQL_Interface(host, user, pwd, db,table_name)
    item.import_file("C:\\Users\\legend\\Desktop\\工作簿1.xlsx")